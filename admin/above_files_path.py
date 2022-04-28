import os

os.system('start osk')

"""
from openpyxl import load_workbook, Workbook
import random
import concurrent.futures

def write():

	wb = load_workbook('Stock.xlsx', read_only=True)
	sheet = wb.active

	WB = Workbook(write_only=True)
	WS = WB.create_sheet()
	WS.append([
			'ref',
			'name',
			"buy",#buy
			'sell',#sell
			'quantity',
			"expired",
			])
	

	for row in range(2, sheet.max_row + 1):

		WS.append([
			sheet.cell(row, 1).value.strip(),
			sheet.cell(row, 4).value.strip(),
			random.randint(4, 5) * row,#buy
			random.randint(1, 3) * row,#sell
			20,
			f'2022-{random.randint(1, 12)}-{random.randint(1, 30)}',
			])

		if row == 100:
			break

		print(row)




	WB.save('file.xlsx')

write()



"""