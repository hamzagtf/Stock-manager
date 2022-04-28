from openpyxl import load_workbook
from sql_db.main import Product_manager
from sql_db.db_sql import Connect_db
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.Qt import Qt
from .excel import Ui_Dialog as AddProduct
from .spinner import Ui_Dialog as Spinner
import concurrent.futures
import datetime
import time
from time import sleep
import threading

class ImportDB:
	def __init__(self, this):
		self.this = this
		self.ui = None
		self.filename = None
		self.val = 0
		


	def importExelData(self):
		#load Excel file
		if self.filename.endswith('xlsx') == False:
			return 'fileNoteSupported'
		try:
			ws = load_workbook(self.filename)
			sheet = ws.active

			if sheet.max_column != 6:
				return 'colError'

			
			self.this.ui.file_link.setText(self.filename)
			table = self.this.ui.excel_tale
			
			table.setRowCount(sheet.max_row - 1)
			row = 0

			for  i in range(2, sheet.max_row + 1):

				#count the precentage
				percentage = (i * 100) / sheet.max_row

				ref = sheet.cell(i, 1).value
				name = sheet.cell(i, 2).value
				buy = sheet.cell(i, 3).value
				sell = sheet.cell(i, 4).value
				quantity = sheet.cell(i, 5).value
				expired = sheet.cell(i, 6).value
				created_at = datetime.datetime.now()

				table.setItem(row, 0, QtWidgets.QTableWidgetItem(str(ref)))
				table.setItem(row, 1, QtWidgets.QTableWidgetItem(str(name)))
				table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(quantity)))
				table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(buy)))
				table.setItem(row, 4, QtWidgets.QTableWidgetItem(str(sell)))
				table.setItem(row, 5, QtWidgets.QTableWidgetItem(str(expired)))
				table.setItem(row, 6, QtWidgets.QTableWidgetItem(str(created_at)))

				row += 1
				self.ui.start.setEnabled(True)

				self.ui.progressBar.setValue(percentage)
				

			self.Dialog.close()

			return True

		except Exception as e:
			raise e
	
	def prograssStates(self, val):
		self.ui.progressBar.setValue(val)
		if self.ui.progressBar.value() == 100:
			self.Dialog.close()

	def prograssTest(self):
		self.threadClass = ThreadClass(self.fetchTableInfo, parent=None,)
		self.threadClass.start()
		self.threadClass.any_signal.connect(self.prograssStates)
		#threadClass.trigger.connect(self.prograssStates)

	

	def spinner(self, *filename):
		

		#progress bar dialog box
		self.Dialog = QtWidgets.QDialog()
		self.ui = Spinner()
		self.ui.setupUi(self.Dialog)
		self.Dialog.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
		self.Dialog.setWindowFlag(Qt.WindowMaximizeButtonHint, False)
		page_number = self.this.ui.stackedWidget.currentIndex()
		
		if page_number == 6:
			self.ui.start.clicked.connect(lambda:self.importExelData())

		else:
			self.ui.start.hide()
			self.prograssTest()

			

		self.Dialog.show()
		

		res = self.Dialog.exec_()
		



	def openFile(self):
		#open file dialog
		filename = QFileDialog.getOpenFileName(self.this, "open File","", "All Files (*);; Excel file (.xlsx)")
		if filename[0]:
			self.filename = filename[0]
		
			self.spinner(filename[0])	
			self.this.ui.stackedWidget.setCurrentWidget(self.this.ui.page_11)

		return False


	def fetchTableInfo(self):
		""" this function return a 
		list inside list containing 
		the all data of  table """

		table = self.this.ui.excel_tale
		row = table.rowCount()
		col = table.columnCount()
		result = []
		
		for r in range(row): 
			data = []
			result.append(data)
			for j in range(col):
				data.append(table.item(r, j).text())

		return result




class ThreadClass(QtCore.QThread):
	any_signal = QtCore.pyqtSignal(float)
	def __init__(self, func,  parent=None):
		super(ThreadClass, self).__init__(parent)
		self.func = func()
		self.CON = Product_manager()
		self.CONN = Connect_db()
	
 

	def run(self):
		
		data = self.func
		
		#data
		#return list inside list
		#return all data inside table
		
		#check product exsitence
		
		for index, i in enumerate(data, 1):
			percentage = (index * 100 ) / len(data)
			ref = i[0]
			name = i[1]
			quantity = i[2]
			buy = i[3]
			sell = i[4]
			expired = i[5]
			created_at = i[6]
			name = name.replace('"', '')
			name = name.replace("'", '')

			CONDITION = f'name="{name}"'	
			checkProductExistence = True if self.CON.fetch_products(CONDITION) else False
			if checkProductExistence:
				#update product if exist
				self.CONN.update_('products', 
					CONDITION,
					ref=ref, name=name, price=sell, quantity=quantity,
					price_buy=buy, expire=expired, created_at=created_at)
			else:
				#create products
				self.CON.createProduct(ref, name, sell, 
								quantity, buy, "none", expired, created_at)

			
			self.any_signal.emit(round(percentage))
			


			

	

	def stop(self):
		print('stop')
		self.terminate()

